"""
Stage 6: Restructure Scoring sheet (5 tech cols -> 9 group cols) and
generate one Q_<id> sheet per candidate.

What this changes
-----------------
Scoring sheet column layout:

  OLD: O Controls/GNC  P Math  Q Coding  R Debug  S Comm
       T Disqualified  U Weighted  V Rank  W Notes
       X UniTier  Y Dept  Z GPA  AA Status  AB Teknofest
       AC Tech Composite  AD DQ Pos

  NEW: O Aircraft Dynamics  P Dynamics Modeling & Sim  Q Guidance
       R Navigation  S Control  T System Identification
       U Company Platforms  V MATLAB/Simulink/Programming
       W ArduPilot Code & Methodology
       X Disqualified?  Y Weighted Score  Z Rank  AA Notes
       AB UniTier Score  AC Dept Score  AD GPA Score
       AE Status Score  AF Teknofest Score
       AG Tech Composite (=AVERAGE of O..W, blanks skipped)
       AH DQ Row Pos

  Columns O..W are FORMULAS pulling the per-group average from each
  candidate's Q_<id> sheet. Editing a Grade on the student sheet
  propagates through to the rank.

Per-student sheet layout (Q_<id>)
---------------------------------
  Row 1: Student name + Tech Composite (live formula).
  Row 2: Free-form general interview notes (merged across full row).
  Row 3: Column headers.
  Rows 4+: 9 group blocks. Each block = 1 banner row (group name +
           live group average in col G) followed by 10 question rows
           referencing QuestionBank for #, Group, Q-EN, Q-TR, A-EN, A-TR.
           Cols G (Grade 1-5) and H (Notes) are interviewer inputs.

The Ranking sheet's INDEX/MATCH formulas are retargeted from Scoring!T/U/V
to Scoring!X/Y/Z to follow the column shift.

Idempotent: re-running drops and recreates each Q_<id> sheet, but
PRESERVES any existing Grade (G) and Notes (H) values + the general-notes
text in row 2 (read into memory before drop, written back after).

Run after stage 5 (QuestionBank must exist with the row index in col I).
"""
import json
import pathlib
import re
import shutil
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

warnings.filterwarnings("ignore")

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX = ROOT / "gnc_june_2026" / "gnc_intern_evaluation_v1.xlsx"
JSON_AUTO = ROOT / "extracted" / "candidates_auto.json"
JSON_OVR = ROOT / "extracted" / "candidates_overrides.json"
JSON_BANK = ROOT / "extracted" / "question_bank.json"

# Row range of candidates on Scoring (matches the existing template)
ROW_FIRST = 3
ROW_LAST = 32

# Number of question groups (must match the JSON bank)
N_GROUPS = 9
N_QPG = 10  # questions per group (uniform; mid-loop divergence OK to handle later)

# Group block geometry on each Q_<id> sheet
QH_TITLE = 1
QH_GEN_NOTES = 2
QH_HEADERS = 3
QH_FIRST_BLOCK = 4   # banner of group 1
# block size = 1 banner + N_QPG questions; e.g. 11 rows per group


# ============ Styles ============
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GROUP_BANNER_FILL = PatternFill("solid", fgColor="305496")
GROUP_BANNER_FONT = Font(bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow for grade/notes
AUTO_FILL = PatternFill("solid", fgColor="E2EFDA")   # green for auto
TITLE_FILL = PatternFill("solid", fgColor="203864")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOPLEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)


# ============ Helpers ============
_TR_MAP = str.maketrans({"ı": "i", "İ": "i", "ş": "s", "Ş": "s",
                          "ğ": "g", "Ğ": "g", "ç": "c", "Ç": "c",
                          "ü": "u", "Ü": "u", "ö": "o", "Ö": "o"})


def deaccent(s: str) -> str:
    import unicodedata
    s = s.translate(_TR_MAP)
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def safe_sheet_name(name: str) -> str:
    """Excel sheet-name rules: <=31 chars, no : \\ / ? * [ ]"""
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name)
    return cleaned[:31]


def quote_sheet_ref(name: str) -> str:
    """Wrap sheet name in single quotes for Excel formula refs, escaping internal quotes."""
    return "'" + name.replace("'", "''") + "'"


def load_candidates_ordered():
    """Reproduce the ordering used by 03_fill_workbook.py so row 3..32 match."""
    auto = json.loads(JSON_AUTO.read_text(encoding="utf-8"))
    overrides = json.loads(JSON_OVR.read_text(encoding="utf-8"))
    overrides.pop("_doc", None)
    merged = []
    for c in auto:
        ovr = overrides.get(c["id"], {})
        merged.append({**c, **{k: v for k, v in ovr.items() if not k.startswith("_")}})
    merged.sort(key=lambda c: (c["name"] or "").lower())
    return merged


def load_question_bank_groups():
    """Return ordered list of {key, display, qb_start, qb_end} from the
    QuestionBank index hidden in column I of stage 5's output."""
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["QuestionBank"]
    groups = []
    r = 2
    while True:
        v = ws.cell(r, 9).value
        if not v:
            break
        key, span = v.split(":")
        s, e = (int(x) for x in span.split("-"))
        # find display from col B at start row
        display = ws.cell(s, 2).value
        groups.append({"key": key, "display": display, "qb_start": s, "qb_end": e})
        r += 1
    wb.close()
    return groups


def read_existing_grades_and_notes(wb, sheet_name, n_groups, n_qpg):
    """If a Q_<id> sheet already exists, pull existing Grade (col G) + Notes
    (col H) values for each question and the general-notes cell B2, so we
    can rewrite without losing them."""
    if sheet_name not in wb.sheetnames:
        return {}, ""
    ws = wb[sheet_name]
    grades = {}
    general_notes = ws.cell(QH_GEN_NOTES, 2).value or ""
    for gi in range(n_groups):
        banner_row = QH_FIRST_BLOCK + gi * (1 + n_qpg)
        for qi in range(n_qpg):
            row = banner_row + 1 + qi
            g = ws.cell(row, 7).value
            n = ws.cell(row, 8).value
            if g is not None or n is not None:
                grades[(gi, qi)] = (g, n)
    return grades, general_notes


def write_student_sheet(wb, sheet_name, candidate_display_name, groups,
                        preserved_grades, preserved_general_notes):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    qb = "QuestionBank"
    qbq = quote_sheet_ref(qb)

    # Column widths
    widths = [5, 28, 50, 50, 60, 60, 10, 35]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # ----- Row 1: title + Tech Composite (live) -----
    ws.cell(QH_TITLE, 1, "Candidate:")
    c = ws.cell(QH_TITLE, 2, candidate_display_name)
    c.font = Font(bold=True, size=12)
    ws.cell(QH_TITLE, 6, "Tech Composite:").font = Font(bold=True)
    # Banner G cells for each group: G4, G(4+11), G(4+22), ...
    banner_gs = [f"G{QH_FIRST_BLOCK + gi * (1 + N_QPG)}" for gi in range(N_GROUPS)]
    composite_formula = "=IFERROR(AVERAGE(" + ",".join(banner_gs) + '),"")'
    cc = ws.cell(QH_TITLE, 7, composite_formula)
    cc.font = Font(bold=True, color="203864")
    cc.alignment = CENTER
    cc.fill = AUTO_FILL
    for col in (1, 6, 7):
        ws.cell(QH_TITLE, col).fill = TITLE_FILL if col != 7 else AUTO_FILL
        if col != 7:
            ws.cell(QH_TITLE, col).font = TITLE_FONT
            ws.cell(QH_TITLE, col).alignment = CENTER
    ws.cell(QH_TITLE, 2).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.row_dimensions[QH_TITLE].height = 24

    # ----- Row 2: general interview notes (merged A:H) -----
    ws.cell(QH_GEN_NOTES, 1, "General Notes:").font = Font(italic=True, color="606060")
    ws.cell(QH_GEN_NOTES, 2, preserved_general_notes)
    ws.merge_cells(start_row=QH_GEN_NOTES, start_column=2,
                   end_row=QH_GEN_NOTES, end_column=8)
    ws.cell(QH_GEN_NOTES, 2).fill = INPUT_FILL
    ws.cell(QH_GEN_NOTES, 2).alignment = TOPLEFT
    ws.row_dimensions[QH_GEN_NOTES].height = 38

    # ----- Row 3: column headers -----
    headers = ["#", "Group", "Question (EN)", "Question (TR)",
               "Reference Answer (EN)", "Reference Answer (TR)",
               "Grade (1-5)", "Notes"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(QH_HEADERS, ci, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[QH_HEADERS].height = 28

    # ----- Group blocks -----
    for gi, g in enumerate(groups):
        banner_row = QH_FIRST_BLOCK + gi * (1 + N_QPG)
        first_q_row = banner_row + 1
        last_q_row = banner_row + N_QPG

        # Group banner — group display name in cols A:F merged; G = group avg; H label
        ws.cell(banner_row, 1, f"{gi+1}.  {g['display']}")
        ws.merge_cells(start_row=banner_row, start_column=1,
                       end_row=banner_row, end_column=6)
        for col in range(1, 7):
            cell = ws.cell(banner_row, col)
            cell.fill = GROUP_BANNER_FILL
            cell.font = GROUP_BANNER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center")
        # Group average
        avg_cell = ws.cell(
            banner_row, 7,
            f'=IFERROR(AVERAGE(G{first_q_row}:G{last_q_row}),"")'
        )
        avg_cell.fill = AUTO_FILL
        avg_cell.font = Font(bold=True, color="203864")
        avg_cell.alignment = CENTER
        avg_cell.border = BORDER
        h_cell = ws.cell(banner_row, 8, "← group avg")
        h_cell.fill = GROUP_BANNER_FILL
        h_cell.font = Font(italic=True, color="FFFFFF", size=10)
        h_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[banner_row].height = 22

        # Question rows — formula-reference the QuestionBank by absolute row
        for qi in range(N_QPG):
            qb_row = g["qb_start"] + qi
            row = first_q_row + qi
            ws.cell(row, 1, f"={qbq}!A{qb_row}")
            ws.cell(row, 2, f"={qbq}!B{qb_row}")
            ws.cell(row, 3, f"={qbq}!C{qb_row}")
            ws.cell(row, 4, f"={qbq}!D{qb_row}")
            ws.cell(row, 5, f"={qbq}!E{qb_row}")
            ws.cell(row, 6, f"={qbq}!F{qb_row}")

            for col in range(1, 7):
                cell = ws.cell(row, col)
                cell.alignment = TOPLEFT
                cell.border = BORDER
                cell.fill = PatternFill("solid", fgColor="F2F2F2")

            # Restore preserved grade + per-question note if present
            g_val, n_val = preserved_grades.get((gi, qi), (None, None))
            gc = ws.cell(row, 7, g_val)
            gc.fill = INPUT_FILL
            gc.alignment = CENTER
            gc.border = BORDER
            gc.font = Font(bold=True)

            nc = ws.cell(row, 8, n_val)
            nc.fill = INPUT_FILL
            nc.alignment = TOPLEFT
            nc.border = BORDER

            ws.row_dimensions[row].height = 60


def build_weighted_score_formula(n):
    """Y<n> weighted score, mirrors the original U formula but uses the new
    column letters (AB..AF auto-component scores, AG tech composite, H..N
    1-5 inputs)."""
    return (
        f'=IF(B{n}="","",('
        f'AB{n}*Weights!$B$2+AC{n}*Weights!$B$3+AD{n}*Weights!$B$4'
        f'+AE{n}*Weights!$B$5+AF{n}*Weights!$B$6+H{n}*Weights!$B$7'
        f'+I{n}*Weights!$B$8+J{n}*Weights!$B$9+K{n}*Weights!$B$10'
        f'+L{n}*Weights!$B$11+M{n}*Weights!$B$12+N{n}*Weights!$B$13'
        f'+AG{n}*Weights!$B$14)/5)'
    )


def build_rank_formula(n):
    """Z<n> sort-stable rank (uses ROW() so it survives sorting)."""
    return (
        f'=IF(OR(B{n}="",X{n}="Y"),"",'
        f'1+SUMPRODUCT(((Y$3:Y$32>Y{n})+'
        f'((Y$3:Y$32=Y{n})*(ROW(Y$3:Y$32)<ROW())))'
        f'*(X$3:X$32<>"Y")*(B$3:B$32<>"")))'
    )


def restructure_scoring(wb, candidates, groups):
    """Read old Scoring values, drop sheet, recreate with new column layout."""
    old_ws = wb["Scoring"]

    # Read the existing candidate data we need to preserve
    rows = []
    for r in range(ROW_FIRST, ROW_LAST + 1):
        rows.append({
            "A": old_ws.cell(r, 1).value,         # #
            "B": old_ws.cell(r, 2).value,         # Name
            "C": old_ws.cell(r, 3).value,         # University
            "D": old_ws.cell(r, 4).value,         # Department
            "E": old_ws.cell(r, 5).value,         # Academic Status
            "F": old_ws.cell(r, 6).value,         # GPA
            "G": old_ws.cell(r, 7).value,         # Teknofest
            "H": old_ws.cell(r, 8).value,         # Real Project
            "I": old_ws.cell(r, 9).value,         # Portfolio
            "J": old_ws.cell(r, 10).value,        # Coursework
            "K": old_ws.cell(r, 11).value,        # Tools
            "L": old_ws.cell(r, 12).value,        # RC/UAV
            "M": old_ws.cell(r, 13).value,        # English
            "N": old_ws.cell(r, 14).value,        # HR Feedback
            # T (Disqualified) and W (Notes) in old layout
            "T_old": old_ws.cell(r, 20).value,
            "W_old": old_ws.cell(r, 23).value,
        })

    # Drop and recreate
    idx = wb.sheetnames.index("Scoring")
    del wb["Scoring"]
    ws = wb.create_sheet("Scoring", index=idx)

    # Column widths (sensible defaults)
    widths = {
        "A": 4, "B": 24, "C": 22, "D": 28, "E": 22, "F": 8,
        "G": 22, "H": 8, "I": 8, "J": 8, "K": 8, "L": 8, "M": 8, "N": 8,
        "O": 12, "P": 16, "Q": 12, "R": 12, "S": 12, "T": 12, "U": 16,
        "V": 18, "W": 18, "X": 12, "Y": 14, "Z": 7, "AA": 32,
        "AB": 10, "AC": 10, "AD": 10, "AE": 10, "AF": 10, "AG": 12, "AH": 8,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B3"

    # Top-banner row 1 (section labels) — same idea as old layout
    banners = [
        (1, 1, "Identity"),
        (3, 7, "Background"),
        (8, 14, "Experience + HR (1-5)"),
        (15, 23, "Tech Interview (per-group avgs, 1-5)"),
        (24, 27, "Result"),
        (28, 34, "Auto-Computed Component Scores"),
    ]
    for start_c, end_c, label in banners:
        cell = ws.cell(1, start_c, label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        if end_c > start_c:
            ws.merge_cells(start_row=1, start_column=start_c,
                           end_row=1, end_column=end_c)

    # Row 2 headers
    h2 = [
        "#", "Name", "University", "Department", "Academic Status",
        "GPA (0-4)", "Teknofest",
        "Real Project (1-5)", "Portfolio (1-5)", "Coursework (1-5)",
        "Tools (1-5)", "RC/UAV (1-5)", "English (1-5)", "HR Feedback (1-5)",
    ] + [g["display"] for g in groups] + [
        "Disqualified?", "Weighted Score", "Rank", "Notes",
        "UniTier Score", "Dept Score", "GPA Score", "Status Score",
        "Teknofest Score", "Tech Composite", "DQ Row Pos",
    ]
    for ci, h in enumerate(h2, start=1):
        cell = ws.cell(2, ci, h)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.font = Font(bold=True, size=10)
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[2].height = 40

    # Body rows
    for ridx, row_data in enumerate(rows):
        n = ROW_FIRST + ridx
        cand = candidates[ridx] if ridx < len(candidates) else None
        if not cand:
            continue
        student_sheet = safe_sheet_name(f"Q_{cand['id']}")
        qref = quote_sheet_ref(student_sheet)

        # A..N - identity / background / experience / HR (copied)
        ws.cell(n, 1, row_data["A"])
        ws.cell(n, 2, row_data["B"])
        ws.cell(n, 3, row_data["C"])
        ws.cell(n, 4, row_data["D"])
        ws.cell(n, 5, row_data["E"])
        ws.cell(n, 6, row_data["F"])
        ws.cell(n, 7, row_data["G"])
        ws.cell(n, 8, row_data["H"])
        ws.cell(n, 9, row_data["I"])
        ws.cell(n, 10, row_data["J"])
        ws.cell(n, 11, row_data["K"])
        ws.cell(n, 12, row_data["L"])
        ws.cell(n, 13, row_data["M"])
        ws.cell(n, 14, row_data["N"])

        # O..W - 9 per-group averages pulled from the student sheet
        for gi in range(N_GROUPS):
            banner_row = QH_FIRST_BLOCK + gi * (1 + N_QPG)
            col = 15 + gi  # 15 = O
            ws.cell(n, col, f'=IFERROR({qref}!G{banner_row},"")')

        # X = Disqualified (was T)
        ws.cell(n, 24, row_data["T_old"] or "N")
        # Y = Weighted Score
        ws.cell(n, 25, build_weighted_score_formula(n))
        # Z = Rank
        ws.cell(n, 26, build_rank_formula(n))
        # AA = Notes (was W)
        ws.cell(n, 27, row_data["W_old"])
        # AB = UniTier Score
        ws.cell(n, 28,
                f'=IFERROR(VLOOKUP(VLOOKUP(C{n},Uni_Tiers!$A$2:$B$24,2,FALSE()),'
                f'Uni_Tiers!$E$2:$F$6,2,FALSE()),IF(C{n}="",1,2))')
        # AC = Dept Score
        ws.cell(n, 29, f'=IFERROR(VLOOKUP(D{n},Dept_Fit!$A$2:$B$15,2,FALSE()),1)')
        # AD = GPA Score
        ws.cell(n, 30,
                f'=IF(F{n}="",1,IF(F{n}>=3.5,5,IF(F{n}>=3,4,IF(F{n}>=2.5,3,IF(F{n}>=2,2,1)))))')
        # AE = Status Score
        ws.cell(n, 31, f'=IFERROR(VLOOKUP(E{n},Status_Map!$A$2:$B$9,2,FALSE()),1)')
        # AF = Teknofest Score
        ws.cell(n, 32, f'=IFERROR(VLOOKUP(G{n},Teknofest_Map!$A$2:$B$6,2,FALSE()),1)')
        # AG = Tech Composite (skip blanks because IFERROR(...,"") returns text)
        ws.cell(n, 33, f'=IFERROR(AVERAGE(O{n}:W{n}),0)')
        # AH = DQ Row Pos
        ws.cell(n, 34, f'=IF(X{n}="Y",ROW()-2,"")')

        # Yellow fill on input cells (H..N, X)
        for col in (8, 9, 10, 11, 12, 13, 14, 24):
            ws.cell(n, col).fill = INPUT_FILL
        # Green fill on auto cells (O..W are interview-derived but auto)
        for col in range(15, 24):
            ws.cell(n, col).fill = AUTO_FILL
        # Green on auto-computed component scores
        for col in (25, 26, 28, 29, 30, 31, 32, 33, 34):
            ws.cell(n, col).fill = AUTO_FILL
        # Borders all
        for col in range(1, 35):
            ws.cell(n, col).border = BORDER
            ws.cell(n, col).alignment = Alignment(vertical="center", wrap_text=True)


def update_ranking_sheet(wb):
    """Retarget Ranking formulas: Scoring!T -> X, U -> Y, V -> Z."""
    ws = wb["Ranking"]
    for r in range(2, ws.max_row + 1):
        for c in range(2, 9):
            cell = ws.cell(r, c)
            v = cell.value
            if isinstance(v, ArrayFormula):
                text = v.text
            elif isinstance(v, str) and v.startswith("="):
                text = v
            else:
                continue
            text = (text
                    .replace("Scoring!V$3:V$32", "Scoring!Z$3:Z$32")
                    .replace("Scoring!U$3:U$32", "Scoring!Y$3:Y$32")
                    .replace("Scoring!T$3:T$32", "Scoring!X$3:X$32"))
            if isinstance(v, ArrayFormula):
                cell.value = ArrayFormula(v.ref, text)
            else:
                cell.value = text


def update_weights_sheet(wb):
    """Update row 14 description to reflect 9 subscores."""
    ws = wb["Weights"]
    ws.cell(14, 3, ("Average of 9 subscores: Aircraft Dynamics, "
                    "Dynamics Modeling & Sim, Guidance, Navigation, "
                    "Control, System ID, Company Platforms, "
                    "MATLAB/Simulink/Programming, ArduPilot Code & Methodology"))


def update_rubric_sheet(wb, groups):
    """Append 9 group rubric rows (or replace any tech rows if present)."""
    ws = wb["Rubric"]
    # Remove any existing rows whose criterion matches old tech categories
    OLD_TECH = {"Controls/GNC", "Math/Physics", "Coding", "Debugging",
                "Communication", "Tech Interview"}
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and any(t.lower() in str(v).lower() for t in OLD_TECH):
            rows_to_delete.append(r)
    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    # Append a divider + 9 group rubric rows at the bottom
    start = ws.max_row + 2
    ws.cell(start, 1, "Tech Interview — 9 question groups").font = Font(bold=True, italic=True)
    rubrics = {
        "Aircraft Dynamics": [
            "Can't name 6 DoF or any mode",
            "Names DoF; vague on stability concepts",
            "Knows DoF + classical modes (phugoid/short-period/dutch-roll)",
            "Comfortable with stability derivatives + linearisation",
            "Fluent: trim, derivatives, modal analysis, CG vs neutral point",
        ],
        "Dynamics Modeling & Sim": [
            "No model-building experience",
            "Knows Simulink basics; no UAV sim built",
            "Has built a small 3-DOF or kinematic sim",
            "Built 6-DOF nonlinear sim; quaternion attitude; sensors",
            "Built + validated 6-DOF vs flight data; MIL/SIL/HIL",
        ],
        "Guidance": [
            "Confuses guidance with control or nav",
            "Knows waypoint following at a high level",
            "Explains L1 or PN at conceptual level",
            "Implements L1/PN/LOS; handles wind & loiter",
            "Designs trajectory + mission state machine; tuned in flight",
        ],
        "Navigation": [
            "No EKF / INS understanding",
            "Knows GPS + IMU exist; vague on fusion",
            "Explains EKF basics; INS mechanisation outline",
            "15-state INS-EKF; magnetometer cal; bias estimation",
            "Tuned EKF on real platform; outlier rejection; multi-frame fluency",
        ],
        "Control": [
            "Knows PID terms only by name",
            "Can tune a single-loop PID by trial",
            "Cascaded inner/outer + gain scheduling concepts",
            "LQR/H-inf basics; anti-windup; phase/gain margins",
            "Designed + flight-tuned multi-axis controller; constraints; saturation",
        ],
        "System Identification": [
            "Unfamiliar with SysID",
            "Knows step response = simple model fit",
            "ARX/OE basics; persistence of excitation",
            "Designs inputs (3-2-1-1/sweep); residual analysis",
            "Grey-box ID of stability derivatives from flight data + CRLB",
        ],
        "Company Platforms": [
            "No exposure",
            "Read public material; knows airframe class",
            "Understands top-level autopilot + sensor suite",
            "Can describe mode hierarchy + failsafes",
            "Hands-on with platform code/test bench; can extend",
        ],
        "MATLAB / Simulink / Programming": [
            "Beginner; can't write a function",
            "Writes scripts; basic Simulink models",
            "Uses functions, structs, vectorisation; can build a Simulink model with buses",
            "S-functions, code generation, embedded targets",
            "Production-quality codegen workflow + git discipline",
        ],
        "ArduPilot Code & Methodology": [
            "Hasn't looked at ArduPilot source",
            "Knows ArduPilot is open source; ran a SITL example",
            "Navigates repo, knows AP_HAL, runs SITL/MAVProxy",
            "Modifies parameter groups, adds RC aux funcs, reads logs",
            "Contributed PR; understands scheduler, EKF3 plumbing, autotest",
        ],
    }
    r = start + 1
    for g in groups:
        display = g["display"]
        anchors = rubrics.get(display)
        if anchors is None:
            anchors = ["", "", "", "", ""]
        ws.cell(r, 1, display).font = Font(bold=True)
        for i, anchor in enumerate(anchors, start=2):
            ws.cell(r, i, anchor).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1


# ============ Main ============
def main():
    wb = openpyxl.load_workbook(XLSX)
    candidates = load_candidates_ordered()
    groups = load_question_bank_groups()
    assert len(groups) == N_GROUPS, f"expected {N_GROUPS} groups, got {len(groups)}"

    # Step 1 — preserve any existing grade/notes data per candidate sheet
    preserved = {}
    for c in candidates:
        sheet_name = safe_sheet_name(f"Q_{c['id']}")
        grades, gnotes = read_existing_grades_and_notes(wb, sheet_name, N_GROUPS, N_QPG)
        preserved[c["id"]] = (grades, gnotes)

    # Step 2 — restructure Scoring sheet (column shift + new formulas)
    restructure_scoring(wb, candidates, groups)

    # Step 3 — write the per-student sheets (drop + recreate)
    for c in candidates:
        sheet_name = safe_sheet_name(f"Q_{c['id']}")
        grades, gnotes = preserved[c["id"]]
        write_student_sheet(wb, sheet_name, c["name"], groups, grades, gnotes)

    # Step 4 — update Ranking sheet references (T->X, U->Y, V->Z)
    update_ranking_sheet(wb)

    # Step 5 — update Weights row 14 description
    update_weights_sheet(wb)

    # Step 6 — update Rubric tech rows
    update_rubric_sheet(wb, groups)

    wb.save(XLSX)
    print(f"Restructured Scoring + built {len(candidates)} Q_<id> sheets.")
    print(f"Updated Ranking, Weights row 14, Rubric tech rows.")


if __name__ == "__main__":
    main()
