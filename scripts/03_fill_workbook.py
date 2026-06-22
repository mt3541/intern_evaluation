"""
Stage 3: Merge candidates_auto.json + candidates_overrides.json and write
the values into gnc_intern_evaluation_v1.xlsx (Scoring sheet, rows 3..N).

Idempotent: re-running overwrites only the inputs cells (yellow inputs).
Tech-interview cells (O–S) and HR Feedback (N) are CLEARED — those must be
filled after the actual interview / HR call.

Original workbook is backed up to gnc_intern_evaluation_v1.xlsx.bak the
first time this is run.
"""
import json, pathlib, shutil, warnings
import openpyxl

warnings.filterwarnings("ignore")

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX = ROOT / "gnc_june_2026" / "gnc_intern_evaluation_v1.xlsx"
BACKUP = XLSX.with_suffix(".xlsx.bak")
AUTO = ROOT / "extracted" / "candidates_auto.json"
OVR = ROOT / "extracted" / "candidates_overrides.json"

# Scoring sheet column map (row 2 headers cross-checked manually)
COL = {
    "name": "B",
    "university": "C",
    "department": "D",
    "academic_status": "E",
    "gpa": "F",
    "teknofest": "G",
    "real_project": "H",
    "portfolio_score": "I",
    "coursework": "J",
    "tools": "K",
    "rc_uav": "L",
    "english": "M",
    "hr_feedback": "N",
    "controls_gnc": "O",
    "math_physics": "P",
    "coding": "Q",
    "debugging": "R",
    "communication": "S",
    "disqualified": "T",
    "notes": "W",
}


def load_candidates():
    auto = json.loads(AUTO.read_text(encoding="utf-8"))
    overrides = json.loads(OVR.read_text(encoding="utf-8"))
    overrides.pop("_doc", None)
    # Merge
    merged = []
    for c in auto:
        ovr = overrides.get(c["id"], {})
        merged_c = {**c, **{k: v for k, v in ovr.items() if not k.startswith("_")}}
        merged.append(merged_c)
    # Sort by name for stable ordering
    merged.sort(key=lambda c: (c["name"] or "").lower())
    return merged


def main():
    if not BACKUP.exists():
        shutil.copy2(XLSX, BACKUP)
        print(f"Backup → {BACKUP.name}")

    candidates = load_candidates()
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Scoring"]

    # Clear any candidate rows beyond what we have (rows 3..32 were pre-populated
    # with example data). Limit to existing template rows.
    template_last_row = 32
    for i, c in enumerate(candidates):
        row = 3 + i
        if row > template_last_row:
            print(f"WARN: more than {template_last_row-2} candidates; row {row} skipped")
            break
        ws[f"A{row}"] = i + 1
        ws[f"{COL['name']}{row}"] = c["name"]
        ws[f"{COL['university']}{row}"] = c["university"]
        ws[f"{COL['department']}{row}"] = c["department"]
        ws[f"{COL['academic_status']}{row}"] = c["academic_status"] or None
        ws[f"{COL['gpa']}{row}"] = (c["gpa"] if c["gpa"] != "" else None)
        ws[f"{COL['teknofest']}{row}"] = c["teknofest"]
        ws[f"{COL['real_project']}{row}"] = c["real_project"]
        ws[f"{COL['portfolio_score']}{row}"] = c["portfolio_score"]
        ws[f"{COL['coursework']}{row}"] = c["coursework"]
        ws[f"{COL['tools']}{row}"] = c["tools"]
        ws[f"{COL['rc_uav']}{row}"] = c["rc_uav"]
        ws[f"{COL['english']}{row}"] = c["english"]
        # Interview / HR — CLEAR (must come from interview)
        for col_key in ("hr_feedback", "controls_gnc", "math_physics",
                        "coding", "debugging", "communication"):
            ws[f"{COL[col_key]}{row}"] = None
        ws[f"{COL['disqualified']}{row}"] = c.get("disqualified") or "N"
        ws[f"{COL['notes']}{row}"] = c.get("notes") or ""

    # Wipe any rows past our candidate count (template had 30 rows of sample data)
    last_filled = 3 + len(candidates) - 1
    for row in range(last_filled + 1, template_last_row + 1):
        for col_letter in COL.values():
            ws[f"{col_letter}{row}"] = None
        ws[f"A{row}"] = None

    wb.save(XLSX)
    print(f"Wrote {len(candidates)} candidates → {XLSX.name}")


if __name__ == "__main__":
    main()
