"""
Stage 5: Build the QuestionBank sheet from extracted/question_bank.json.

Reads the JSON, drops any existing QuestionBank sheet, recreates it with
one row per question (group order preserved). Per-student sheets in stage 6
reference this sheet by row, so the ROW ORDER here is the contract — do not
change it without rerunning stage 6.

Backs up the workbook to .bak2 on first run (preserves the earlier .bak).
"""
import json, pathlib, shutil, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")

ROOT = pathlib.Path(__file__).resolve().parent.parent
XLSX = ROOT / "gnc_june_2026" / "gnc_intern_evaluation_v1.xlsx"
BAK2 = XLSX.with_suffix(".xlsx.bak2")
JSON_PATH = ROOT / "extracted" / "question_bank.json"

SHEET = "QuestionBank"
HEADERS = ["#", "Group", "Question (EN)", "Question (TR)",
           "Reference Answer (EN)", "Reference Answer (TR)", "Notes"]
COL_WIDTHS = [5, 30, 50, 50, 60, 60, 30]


def main():
    if not BAK2.exists():
        shutil.copy2(XLSX, BAK2)
        print(f"Backup -> {BAK2.name}")

    bank = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(XLSX)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    # Header row
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(*[Side(style="thin", color="BFBFBF")] * 4)
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for c, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "B2"

    # Group color stripes (alternate light fills per group)
    group_fills = [
        PatternFill("solid", fgColor="EAF1FB"),
        PatternFill("solid", fgColor="FFF5E6"),
        PatternFill("solid", fgColor="ECF7EC"),
        PatternFill("solid", fgColor="FBEFEF"),
        PatternFill("solid", fgColor="F5ECFB"),
        PatternFill("solid", fgColor="ECF5F5"),
        PatternFill("solid", fgColor="FFFBE6"),
        PatternFill("solid", fgColor="EFECFB"),
        PatternFill("solid", fgColor="FBECF1"),
    ]

    row = 2
    group_rows = {}  # group_key -> (start_row, end_row) for stage 6
    for gi, group in enumerate(bank["groups"]):
        key = group["key"]
        display = group["display"]
        start = row
        for qi, q in enumerate(group["questions"], start=1):
            ws.cell(row, 1, qi)
            ws.cell(row, 2, display)
            ws.cell(row, 3, q["q_en"])
            ws.cell(row, 4, q["q_tr"])
            ws.cell(row, 5, q["a_en"])
            ws.cell(row, 6, q["a_tr"])
            ws.cell(row, 7, q.get("notes", "") or "")
            for c in range(1, 8):
                cell = ws.cell(row, c)
                cell.fill = group_fills[gi % len(group_fills)]
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
            ws.row_dimensions[row].height = 60
            row += 1
        end = row - 1
        group_rows[key] = (start, end)

    # Drop a small index in cells past column H so stage 6 can read it back
    # (cheaper than re-parsing JSON in two places). Hidden-ish offsets in col I.
    ws.cell(1, 9, "_GROUP_ROW_INDEX")
    ws.cell(1, 9).font = Font(italic=True, color="808080", size=9)
    for i, (k, (s, e)) in enumerate(group_rows.items(), start=2):
        ws.cell(i, 9, f"{k}:{s}-{e}")
        ws.cell(i, 9).font = Font(color="808080", size=9)
    ws.column_dimensions["I"].hidden = True

    wb.save(XLSX)
    total = row - 2
    print(f"Wrote {total} questions across {len(bank['groups'])} groups -> {SHEET}")
    for k, (s, e) in group_rows.items():
        print(f"  {k:32s}  rows {s}-{e}")


if __name__ == "__main__":
    main()
